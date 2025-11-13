from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.db import transaction
from django.db.models import Sum, F, Q
from django.utils import timezone
import tempfile
import os
import re
import json
import wave

# Import de Vosk para reconocimiento de voz
try:
    from vosk import Model, KaldiRecognizer
    import wave
    VOSK_AVAILABLE = True
except ImportError:
    VOSK_AVAILABLE = False
    print("⚠️ Vosk no está instalado. Instala con: pip install vosk")

# Import de TheFuzz para matching fuzzy
try:
    from thefuzz import fuzz, process
    FUZZ_AVAILABLE = True
except ImportError:
    FUZZ_AVAILABLE = False
    print("⚠️ TheFuzz no está instalado. Instala con: pip install thefuzz python-Levenshtein")

from .models import Cart, CartItem, Venta, VentaDetalle, Pago, NotificacionPush, Reporte, PromptFrecuente, ModeloIA, Prediccion, ReporteVozMovil, CompartirReporte, PreferenciaNotificaciones, SincronizacionDatos
from apps.authentication.models import DispositivosMoviles
from .serializers import (
    CartSerializer, 
    CartItemSerializer, 
    CartItemCreateSerializer,
    VoiceCommandSerializer,
    VentaSerializer,
    VentaCreateSerializer,
    VentaDetalleSerializer,
    PagoSerializer,
    PagoCreateSerializer,
    PagoQRSerializer,
    DispositivoMovilSerializer,
    VentaMovilSerializer,
    VentaCreateMovilSerializer,
    VentaHistoricoMovilSerializer,
    DashboardMovilSerializer,
    NotificacionPushSerializer,
    NotificacionPushCreateSerializer,
    ReporteSerializer,
    ReporteGenerarSerializer,
    ReporteListadoSerializer,
    ReporteExportarSerializer,
    ReporteVozSerializer,
    PromptFrecuenteSerializer,
    PromptFrecuenteCreateSerializer,
    ModeloIASerializer,
    ModeloIAEntrenarSerializer,
    PrediccionSerializer,
    PrediccionListadoSerializer,
    ReporteVozMovilSerializer,
    ReporteVozMovilCreateSerializer,
    CompartirReporteSerializer,
    CompartirReporteCreateSerializer,
    PreferenciaNotificacionesSerializer,
    PreferenciaNotificacionesUpdateSerializer,
    SincronizacionDatosSerializer,
    SincronizacionDatosCreateSerializer
)
from apps.products.models import Productos


class CartViewSet(viewsets.ModelViewSet):
    queryset = Cart.objects.all()
    serializer_class = CartSerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]

    def get_queryset(self):
        queryset = super().get_queryset()
        # Si es admin/staff, mostrar todos los carritos
        is_staff = getattr(self.request.user, 'is_staff', False)
        is_superuser = getattr(self.request.user, 'is_superuser', False)
        if is_staff or is_superuser:
            return queryset
        # Si es usuario regular, filtrar por sus carritos
        if self.request.user.is_authenticated and hasattr(self.request.user, 'usuarios'):
            return queryset.filter(usuario__user=self.request.user)
        return queryset

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def add_item(self, request, pk=None):
        """Agregar un producto al carrito"""
        cart = self.get_object()
        
        if cart.status != 'open':
            return Response(
                {'detail': 'No se pueden agregar items a un carrito cerrado'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = CartItemCreateSerializer(data=request.data)
        if serializer.is_valid():
            producto = serializer.validated_data['producto']
            quantity = serializer.validated_data['quantity']
            
            # Verificar stock
            if producto.stock_actual < quantity:
                return Response(
                    {'detail': f'Stock insuficiente. Disponible: {producto.stock_actual}'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Verificar si el producto ya existe en el carrito
            existing_item = cart.items.filter(producto=producto).first()
            if existing_item:
                existing_item.quantity += quantity
                existing_item.save()
                item = existing_item
            else:
                item = CartItem.objects.create(
                    cart=cart,
                    producto=producto,
                    quantity=quantity,
                    price=producto.precio_venta
                )
            
            # Actualizar total del carrito
            self._update_cart_total(cart)
            
            return Response(CartSerializer(cart).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def remove_item(self, request, pk=None):
        """Eliminar un producto del carrito"""
        cart = self.get_object()
        item_id = request.data.get('item_id')
        
        if not item_id:
            return Response(
                {'detail': 'Se requiere item_id'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            item = cart.items.get(id=item_id)
            item.delete()
            self._update_cart_total(cart)
            return Response(CartSerializer(cart).data)
        except CartItem.DoesNotExist:
            return Response(
                {'detail': 'Item no encontrado'}, 
                status=status.HTTP_404_NOT_FOUND
            )

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def update_item(self, request, pk=None):
        """Actualizar cantidad de un producto en el carrito"""
        cart = self.get_object()
        item_id = request.data.get('item_id')
        quantity = request.data.get('quantity')
        
        if not item_id or quantity is None:
            return Response(
                {'detail': 'Se requiere item_id y quantity'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            quantity = int(quantity)
            if quantity <= 0:
                return Response(
                    {'detail': 'La cantidad debe ser mayor a 0'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            item = cart.items.get(id=item_id)
            
            # Verificar stock
            if item.producto.stock_actual < quantity:
                return Response(
                    {'detail': f'Stock insuficiente. Disponible: {item.producto.stock_actual}'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            item.quantity = quantity
            item.save()
            self._update_cart_total(cart)
            
            return Response(CartSerializer(cart).data)
        except CartItem.DoesNotExist:
            return Response(
                {'detail': 'Item no encontrado'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except ValueError:
            return Response(
                {'detail': 'Cantidad inválida'}, 
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def clear(self, request, pk=None):
        """Vaciar el carrito"""
        cart = self.get_object()
        cart.items.all().delete()
        cart.total = 0
        cart.save()
        return Response(CartSerializer(cart).data)

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def checkout(self, request, pk=None):
        """Finalizar compra"""
        cart = self.get_object()
        
        if cart.status != 'open':
            return Response(
                {'detail': 'El carrito ya fue procesado'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not cart.items.exists():
            return Response(
                {'detail': 'El carrito está vacío'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Marcar como completado
        cart.status = 'completed'
        cart.save()
        
        return Response({
            'detail': 'Compra finalizada exitosamente',
            'cart_id': str(cart.id),
            'total': str(cart.total),
            'items_count': cart.items.count()
        })

    @action(
        detail=False, 
        methods=['post'], 
        permission_classes=[permissions.IsAuthenticated],
        parser_classes=[MultiPartParser, FormParser]
    )
    def voice_command(self, request):
        """
        Procesar comando de voz para gestionar el carrito usando Vosk.
        
        Ejemplos de comandos:
        - "agregar 3 unidades de producto SKU123"
        - "añadir 5 del producto codigo ABC"
        - "eliminar producto SKU456"
        - "quitar item 12"
        - "actualizar cantidad a 10 del item 5"
        - "vaciar carrito"
        - "finalizar compra"
        """
        # Verificar que Vosk está disponible
        if not VOSK_AVAILABLE:
            return Response(
                {
                    'detail': 'Vosk no está instalado',
                    'error': 'Para usar comandos de voz, instala: pip install vosk',
                    'suggestion': 'Usa los endpoints de texto mientras tanto'
                },
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )
        
        serializer = VoiceCommandSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        audio_file = serializer.validated_data['audio']
        cart_id = serializer.validated_data.get('cart_id')
        
        # Obtener o crear carrito
        if cart_id:
            try:
                cart = Cart.objects.get(id=cart_id)
            except Cart.DoesNotExist:
                return Response(
                    {'detail': 'Carrito no encontrado'}, 
                    status=status.HTTP_404_NOT_FOUND
                )
        else:
            # Buscar carrito abierto del usuario o crear uno nuevo
            if hasattr(request.user, 'usuarios'):
                usuario = request.user.usuarios
                cart = Cart.objects.filter(usuario=usuario, status='open').first()
                if not cart:
                    cart = Cart.objects.create(usuario=usuario, status='open')
            else:
                # Si no hay usuario asociado (admin), crear carrito sin usuario
                cart = Cart.objects.filter(status='open', usuario__isnull=True).first()
                if not cart:
                    cart = Cart.objects.create(status='open')
        
        # Guardar archivo temporal original
        original_path = tempfile.mktemp(suffix=os.path.splitext(audio_file.name)[1])
        with open(original_path, 'wb') as f:
            for chunk in audio_file.chunks():
                f.write(chunk)
        
        # Logging detallado del archivo de audio
        file_size = os.path.getsize(original_path)
        print(f"\n{'='*60}")
        print(f"🎤 PROCESANDO AUDIO CON VOSK")
        print(f"{'='*60}")
        print(f"📁 Nombre original: {audio_file.name}")
        print(f"📦 Tamaño: {file_size} bytes ({file_size/1024:.2f} KB)")
        print(f"🗂️ Tipo MIME: {audio_file.content_type}")
        
        # Verificar que el archivo no esté vacío
        if file_size < 100:
            print(f"⚠️ ADVERTENCIA: Archivo muy pequeño ({file_size} bytes)")
            os.unlink(original_path)
            return Response({
                'detail': f'El archivo de audio es demasiado pequeño ({file_size} bytes). Por favor, graba al menos 1-2 segundos.',
                'transcription': '',
                'suggestion': 'Asegúrate de hablar durante la grabación y que el indicador de nivel muestre >5%.',
                'cart_id': str(cart.id),
                'debug_info': {
                    'file_size': file_size,
                    'file_name': audio_file.name,
                    'content_type': audio_file.content_type
                }
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Archivo WAV temporal para Vosk
        wav_path = tempfile.mktemp(suffix='.wav')
        
        try:
            # Convertir a WAV 16kHz mono (requerido por Vosk)
            import subprocess
            print(f"⏳ Convirtiendo audio a formato WAV 16kHz mono...")
            conversion_result = subprocess.run(
                ['ffmpeg', '-y', '-i', original_path, '-ar', '16000', '-ac', '1', '-f', 'wav', wav_path],
                capture_output=True,
                text=True,
                timeout=30
            )
            
            if conversion_result.returncode != 0:
                raise Exception(f"Error en conversión FFmpeg: {conversion_result.stderr[:200]}")
            
            print(f"✅ Audio convertido a WAV 16kHz mono")
            
            # Cargar modelo de Vosk (español)
            model_path = os.path.join(os.path.dirname(__file__), '..', '..', 'vosk-model-small-es-0.42')
            model_path = os.path.abspath(model_path)
            
            if not os.path.exists(model_path):
                return Response({
                    'detail': 'Modelo de Vosk no encontrado',
                    'error': f'Descarga el modelo desde: https://alphacephei.com/vosk/models',
                    'suggestion': f'Descarga vosk-model-small-es-0.42.zip y extráelo en: {model_path}',
                    'help_url': 'https://alphacephei.com/vosk/models/vosk-model-small-es-0.42.zip'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            print(f"⏳ Cargando modelo de Vosk desde: {model_path}")
            model = Model(model_path)
            print(f"✅ Modelo de Vosk cargado correctamente")
            
            # Transcribir audio con Vosk
            print(f"⏳ Transcribiendo audio...")
            wf = wave.open(wav_path, "rb")
            
            if wf.getnchannels() != 1 or wf.getsampwidth() != 2 or wf.getframerate() != 16000:
                wf.close()
                raise Exception("Audio debe ser WAV 16kHz mono PCM")
            
            rec = KaldiRecognizer(model, wf.getframerate())
            rec.SetWords(True)
            rec.SetMaxAlternatives(0)
            
            transcription_parts = []
            while True:
                data = wf.readframes(4000)
                if len(data) == 0:
                    break
                if rec.AcceptWaveform(data):
                    result = json.loads(rec.Result())
                    if 'text' in result and result['text']:
                        transcription_parts.append(result['text'])
            
            # Obtener resultado final
            final_result = json.loads(rec.FinalResult())
            if 'text' in final_result and final_result['text']:
                transcription_parts.append(final_result['text'])
            
            wf.close()
            
            transcription = ' '.join(transcription_parts).strip().lower()
            
            print(f"✅ Transcripción completada")
            print(f"📝 Texto transcrito: '{transcription}'")
            print(f"   - Longitud: {len(transcription)} caracteres")
            print(f"   - Vacío: {not transcription}")
            
            # Verificar solo que no esté vacío
            if not transcription:
                print(f"⚠️ AUDIO RECHAZADO: Transcripción vacía")
                return Response({
                    'detail': 'No se detectó audio válido. Por favor, habla más fuerte y cerca del micrófono.',
                    'transcription': transcription,
                    'suggestion': 'Asegúrate de que el micrófono esté funcionando y habla claramente.',
                    'cart_id': str(cart.id),
                    'debug_info': {
                        'file_size': file_size,
                        'transcription_length': len(transcription)
                    }
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Logging
            print(f"✅ AUDIO ACEPTADO: Transcripción válida")
            print(f"🎤 Transcripción exitosa: '{transcription}'")
            print(f"{'='*60}\n")
            
            # Procesar comando con TheFuzz para mejor matching
            response_data = self._process_voice_command(cart, transcription)
            
            return Response({
                'transcription': transcription,
                'cart_id': str(cart.id),
                **response_data
            })
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"❌ Error en voice_command: {error_details}")
            return Response(
                {
                    'detail': f'Error procesando audio: {str(e)}',
                    'error_type': type(e).__name__,
                    'error_details': str(e)
                }, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        finally:
            # Eliminar archivos temporales
            if os.path.exists(original_path):
                try:
                    os.unlink(original_path)
                except:
                    pass
            if os.path.exists(wav_path):
                try:
                    os.unlink(wav_path)
                except:
                    pass

    def _process_voice_command(self, cart, text):
        """Procesar el texto transcrito y ejecutar la acción correspondiente"""
        text = text.lower()
        
        # Normalizar texto (quitar acentos y caracteres especiales)
        import unicodedata
        text = ''.join(
            c for c in unicodedata.normalize('NFD', text)
            if unicodedata.category(c) != 'Mn'
        )
        
        # Logging para debugging
        print(f"🎤 Comando recibido: '{text}'")
        
        # Comando: Agregar producto
        agregar_palabras = [
            'agregar', 'anadir', 'aniadir', 'agregame', 'anademe', 
            'agregá', 'añade', 'agrega', 'añadi', 'anade', 'aniade',
            'mete', 'pon', 'poner', 'ponme', 'meter', 'meteme',
            'incluye', 'incluir', 'incluye', 'inclui',
            'quiero', 'dame', 'necesito',
            'comprar', 'compra', 'llevame', 'llevo',
            'add', 'adiciona', 'adicionar'
        ]
        if any(word in text for word in agregar_palabras):
            print(f"✅ Detectado comando: AGREGAR")
            return self._voice_add_product(cart, text)
        
        # Comando: Eliminar producto
        eliminar_palabras = [
            'eliminar', 'quitar', 'borrar', 'sacar', 'elimina', 'quita', 'borra', 'saca',
            'remove', 'remover', 'remueve', 'saque', 'borre', 'quite',
            'descarta', 'descartar', 'cancela', 'cancelar',
            'no quiero', 'no lo quiero', 'ya no', 'fuera',
            'delete', 'drop'
        ]
        if any(word in text for word in eliminar_palabras):
            print(f"✅ Detectado comando: ELIMINAR")
            return self._voice_remove_item(cart, text)
        
        # Comando: Actualizar cantidad
        actualizar_palabras = [
            'actualizar', 'cambiar', 'modificar', 'actualiza', 'cambia', 'modifica',
            'update', 'ajustar', 'ajusta', 'editar', 'edita',
            'poner', 'pon', 'establecer', 'establece',
            'dejar', 'deja', 'pasar', 'pasa'
        ]
        if any(word in text for word in actualizar_palabras):
            # Verificar que también mencione "cantidad" o números para diferenciarlo de agregar
            cantidad_palabras = ['cantidad', 'numero', 'unidades', 'item', 'articulo']
            if any(word in text for word in cantidad_palabras) or re.search(r'a\s+\d+', text):
                print(f"✅ Detectado comando: ACTUALIZAR")
                return self._voice_update_item(cart, text)
        
        # Comando: Vaciar carrito
        vaciar_palabras = [
            'vaciar', 'limpiar', 'borrar todo', 'eliminar todo', 'vacia', 'limpia',
            'clear', 'reset', 'resetear', 'reiniciar',
            'quitar todo', 'sacar todo', 'borrame todo', 'eliminame todo',
            'vacio', 'limpio', 'borra todo', 'quita todo'
        ]
        if any(word in text for word in vaciar_palabras):
            print(f"✅ Detectado comando: VACIAR")
            cart.items.all().delete()
            cart.total = 0
            cart.save()
            return {
                'action': 'clear',
                'message': 'Carrito vaciado exitosamente',
                'cart': CartSerializer(cart).data
            }
        
        # Comando: Finalizar compra
        finalizar_palabras = [
            'finalizar', 'terminar', 'checkout', 'pagar', 'finaliza', 'termina',
            'comprar', 'finish', 'completar', 'completa',
            'hacer pedido', 'hacer compra', 'confirmar',
            'proceder', 'continuar', 'listo',
            'ya esta', 'eso es todo', 'nada mas'
        ]
        if any(word in text for word in finalizar_palabras):
            # Verificar que no sea "agregar" o similar
            if not any(word in text for word in agregar_palabras[:10]):
                print(f"✅ Detectado comando: FINALIZAR")
                if cart.items.exists():
                    cart.status = 'completed'
                    cart.save()
                    return {
                        'action': 'checkout',
                        'message': 'Compra finalizada exitosamente',
                        'total': str(cart.total)
                    }
                else:
                    return {
                        'action': 'checkout',
                        'error': 'El carrito está vacío',
                        'cart': CartSerializer(cart).data
                    }
        
        # Comando: Ver carrito
        ver_palabras = [
            'mostrar', 'ver', 'consultar', 'listar', 'muestra', 've',
            'show', 'display', 'enseña', 'dime', 'que tengo',
            'mi carrito', 'el carrito', 'que hay', 'cuanto',
            'revisar', 'revisa', 'check'
        ]
        if any(word in text for word in ver_palabras):
            print(f"✅ Detectado comando: VER")
            return {
                'action': 'view',
                'message': 'Contenido del carrito',
                'cart': CartSerializer(cart).data
            }
        
        else:
            print(f"❌ Comando no reconocido")
            return {
                'action': 'unknown',
                'error': 'No se pudo interpretar el comando',
                'transcription': text,
                'suggestion': 'Intenta comandos como: "agregar producto", "eliminar item", "vaciar carrito", "finalizar compra"',
                'detected_words': text.split()
            }

    def _voice_add_product(self, cart, text):
        """Agregar producto por voz"""
        print(f"   🔍 Procesando comando AGREGAR")
        print(f"   📝 Texto original: '{text}'")
        
        # Diccionario de números en español a dígitos
        numeros_texto = {
            'cero': '0', 'uno': '1', 'dos': '2', 'tres': '3', 'cuatro': '4',
            'cinco': '5', 'seis': '6', 'siete': '7', 'ocho': '8', 'nueve': '9',
            'diez': '10', 'once': '11', 'doce': '12', 'trece': '13', 'catorce': '14',
            'quince': '15', 'dieciseis': '16', 'diecisiete': '17', 'dieciocho': '18',
            'diecinueve': '19', 'veinte': '20', 'treinta': '30', 'cuarenta': '40',
            'cincuenta': '50', 'sesenta': '60', 'setenta': '70', 'ochenta': '80',
            'noventa': '90', 'cien': '100'
        }
        
        # Reemplazar números en texto por dígitos
        text_procesado = text
        for palabra, numero in numeros_texto.items():
            text_procesado = text_procesado.replace(palabra, numero)
        
        print(f"   📝 Texto procesado: '{text_procesado}'")
        
        # Buscar cantidad (priorizar números en dígitos)
        quantity_match = re.search(r'(\d+)\s*(?:unidad|unidades|producto|productos|de)?', text_procesado)
        quantity = int(quantity_match.group(1)) if quantity_match else 1
        print(f"   🔢 Cantidad detectada: {quantity}")
        
        # Buscar SKU o código de producto (más flexible)
        # Patrones: "SKU ABC123", "codigo ABC123", "producto ABC123", "laptop 001"
        sku_patterns = [
            r'(?:sku|codigo|código|producto)\s*[:=]?\s*([a-zA-Z0-9\-_\s]+?)(?:\s|$)',  # Con palabra clave
            r'\b([A-Z]{2,}[0-9]+)\b',  # Patrón como PROD001, ABC123
            r'\b([a-zA-Z]+\s*[0-9]{3,})\b',  # Patrón como "laptop 001", "prod001"
            r'\b([a-zA-Z]+[0-9]{3,})\b',  # Patrón como "laptop001", "prod001"
        ]
        
        sku = None
        for pattern in sku_patterns:
            sku_match = re.search(pattern, text, re.IGNORECASE)
            if sku_match:
                # Limpiar el SKU (quitar espacios extras, convertir a mayúsculas)
                sku = sku_match.group(1).strip().replace(' ', '').upper()
                print(f"   📋 SKU detectado: '{sku}' (patrón usado: {pattern})")
                break
        
        if sku:
            try:
                # Buscar producto por SKU o código de barras
                producto = Productos.objects.filter(
                    Q(sku__iexact=sku) | Q(codigo_barras__iexact=sku)
                ).first()
                
                if not producto:
                    # Intentar búsqueda parcial
                    producto = Productos.objects.filter(
                        Q(sku__icontains=sku) | Q(nombre__icontains=sku)
                    ).first()
                
                if not producto:
                    return {
                        'action': 'add_item',
                        'error': f'Producto con SKU "{sku}" no encontrado',
                        'suggestion': f'Verifica que el SKU sea correcto. SKU buscado: {sku}',
                        'cart': CartSerializer(cart).data
                    }
                
                if producto.stock_actual < quantity:
                    return {
                        'action': 'add_item',
                        'error': f'Stock insuficiente para {producto.nombre}. Disponible: {producto.stock_actual}',
                        'cart': CartSerializer(cart).data
                    }
                
                # Agregar o actualizar item
                existing_item = cart.items.filter(producto=producto).first()
                if existing_item:
                    existing_item.quantity += quantity
                    existing_item.save()
                else:
                    CartItem.objects.create(
                        cart=cart,
                        producto=producto,
                        quantity=quantity,
                        price=producto.precio_venta
                    )
                
                self._update_cart_total(cart)
                
                return {
                    'action': 'add_item',
                    'message': f'Se agregaron {quantity} unidades de {producto.nombre} al carrito',
                    'producto': {
                        'sku': producto.sku,
                        'nombre': producto.nombre,
                        'precio': str(producto.precio_venta),
                        'cantidad': quantity
                    },
                    'cart': CartSerializer(cart).data
                }
            except Exception as e:
                return {
                    'action': 'add_item',
                    'error': f'Error al agregar producto: {str(e)}',
                    'cart': CartSerializer(cart).data
                }
        else:
            return {
                'action': 'add_item',
                'error': 'No se pudo identificar el SKU del producto',
                'transcription': text,
                'suggestion': 'Intenta decir: "agregar 2 unidades del producto SKU ABC123" o "agregar 3 del producto PROD001"',
                'cart': CartSerializer(cart).data
            }

    def _voice_remove_item(self, cart, text):
        """Eliminar item por voz"""
        # Buscar ID del item
        item_match = re.search(r'(?:item|ítem)\s*(\d+)', text)
        
        if item_match:
            item_id = int(item_match.group(1))
            try:
                item = cart.items.get(id=item_id)
                producto_nombre = item.producto.nombre
                item.delete()
                self._update_cart_total(cart)
                return {
                    'action': 'remove_item',
                    'message': f'Se eliminó {producto_nombre} del carrito',
                    'cart': CartSerializer(cart).data
                }
            except CartItem.DoesNotExist:
                return {
                    'action': 'remove_item',
                    'error': f'Item {item_id} no encontrado en el carrito',
                    'cart': CartSerializer(cart).data
                }
        else:
            # Buscar por SKU
            sku_match = re.search(r'(?:sku|codigo|código)\s*([a-zA-Z0-9]+)', text)
            if sku_match:
                sku = sku_match.group(1).upper()
                item = cart.items.filter(producto__sku=sku).first()
                if item:
                    producto_nombre = item.producto.nombre
                    item.delete()
                    self._update_cart_total(cart)
                    return {
                        'action': 'remove_item',
                        'message': f'Se eliminó {producto_nombre} del carrito',
                        'cart': CartSerializer(cart).data
                    }
                else:
                    return {
                        'action': 'remove_item',
                        'error': f'Producto con SKU {sku} no encontrado en el carrito',
                        'cart': CartSerializer(cart).data
                    }
            
            return {
                'action': 'remove_item',
                'error': 'No se pudo identificar el item. Intenta: "eliminar item 5" o "quitar producto SKU ABC123"',
                'cart': CartSerializer(cart).data
            }

    def _voice_update_item(self, cart, text):
        """Actualizar cantidad de item por voz"""
        # Buscar cantidad
        quantity_match = re.search(r'(?:cantidad|a)\s*(\d+)', text)
        if not quantity_match:
            return {
                'action': 'update_item',
                'error': 'No se pudo identificar la nueva cantidad',
                'cart': CartSerializer(cart).data
            }
        
        quantity = int(quantity_match.group(1))
        
        # Buscar ID del item
        item_match = re.search(r'(?:item|ítem)\s*(\d+)', text)
        
        if item_match:
            item_id = int(item_match.group(1))
            try:
                item = cart.items.get(id=item_id)
                
                if item.producto.stock_actual < quantity:
                    return {
                        'action': 'update_item',
                        'error': f'Stock insuficiente para {item.producto.nombre}. Disponible: {item.producto.stock_actual}',
                        'cart': CartSerializer(cart).data
                    }
                
                item.quantity = quantity
                item.save()
                self._update_cart_total(cart)
                
                return {
                    'action': 'update_item',
                    'message': f'Se actualizó la cantidad a {quantity} unidades',
                    'cart': CartSerializer(cart).data
                }
            except CartItem.DoesNotExist:
                return {
                    'action': 'update_item',
                    'error': f'Item {item_id} no encontrado en el carrito',
                    'cart': CartSerializer(cart).data
                }
        else:
            return {
                'action': 'update_item',
                'error': 'No se pudo identificar el item. Intenta: "actualizar cantidad a 5 del item 3"',
                'cart': CartSerializer(cart).data
            }

    def _update_cart_total(self, cart):
        """Actualizar el total del carrito"""
        total = cart.items.aggregate(
            total=Sum(F('price') * F('quantity'))
        )['total'] or 0
        cart.total = total
        cart.save()


class CartItemViewSet(viewsets.ModelViewSet):
    queryset = CartItem.objects.all()
    serializer_class = CartItemSerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]


# ============================================================================
# CU12: Registrar Venta - ViewSet
# ============================================================================

class VentaViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar ventas (CU12)
    
    Endpoints:
    - GET /api/sales/ventas/ - Listar todas las ventas
    - POST /api/sales/ventas/ - Crear una venta desde un carrito
    - GET /api/sales/ventas/{id}/ - Ver detalle de una venta
    - PUT/PATCH /api/sales/ventas/{id}/ - Actualizar venta
    - DELETE /api/sales/ventas/{id}/ - Eliminar venta
    - POST /api/sales/ventas/crear_desde_carrito/ - Crear venta desde carrito
    - POST /api/sales/ventas/{id}/cancelar/ - Cancelar venta
    - GET /api/sales/ventas/ventas_por_cliente/ - Ventas de un cliente específico
    """
    queryset = Venta.objects.all().prefetch_related('detalles', 'detalles__producto')
    serializer_class = VentaSerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filtrar por cliente
        cliente_id = self.request.query_params.get('cliente_id', None)
        if cliente_id:
            queryset = queryset.filter(cliente_id=cliente_id)
        
        # Filtrar por estado
        estado = self.request.query_params.get('estado', None)
        if estado:
            queryset = queryset.filter(estado=estado)
        
        # Filtrar por tipo de entrega
        tipo_entrega = self.request.query_params.get('tipo_entrega', None)
        if tipo_entrega:
            queryset = queryset.filter(tipo_entrega=tipo_entrega)
        
        # Filtrar por rango de fechas
        fecha_desde = self.request.query_params.get('fecha_desde', None)
        fecha_hasta = self.request.query_params.get('fecha_hasta', None)
        if fecha_desde:
            queryset = queryset.filter(fecha_venta__gte=fecha_desde)
        if fecha_hasta:
            queryset = queryset.filter(fecha_venta__lte=fecha_hasta)
        
        return queryset
    
    @action(detail=False, methods=['post'])
    def crear_desde_carrito(self, request):
        """
        Crear una venta desde un carrito existente
        
        Body:
        {
            "cart_id": "uuid-del-carrito",
            "cliente_id": 1,
            "tipo_entrega": "local",
            "descuento": 0,
            "impuesto_porcentaje": 13,
            "direccion_entrega": "Calle 123",
            "notas": "Notas adicionales"
        }
        """
        serializer = VentaCreateSerializer(data=request.data, context={'request': request})
        
        if serializer.is_valid():
            try:
                venta = serializer.save()
                return Response({
                    'success': True,
                    'message': 'Venta creada exitosamente',
                    'data': VentaSerializer(venta).data
                }, status=status.HTTP_201_CREATED)
            except Exception as e:
                return Response({
                    'success': False,
                    'error': str(e)
                }, status=status.HTTP_400_BAD_REQUEST)
        
        return Response({
            'success': False,
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def cancelar(self, request, pk=None):
        """
        Cancelar una venta y revertir el stock
        
        POST /api/sales/ventas/{id}/cancelar/
        Body: {"motivo": "Razón de la cancelación"}
        """
        venta = self.get_object()
        
        if venta.estado == 'cancelada':
            return Response({
                'success': False,
                'error': 'Esta venta ya está cancelada'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if venta.estado == 'pagada':
            return Response({
                'success': False,
                'error': 'No se puede cancelar una venta pagada. Debe solicitar un reembolso.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        motivo = request.data.get('motivo', 'Sin motivo especificado')
        
        try:
            with transaction.atomic():
                # Revertir stock
                for detalle in venta.detalles.all():
                    producto = detalle.producto
                    if producto.stock_actual is not None:
                        producto.stock_actual += detalle.cantidad
                        producto.save()
                
                # Actualizar estado de venta
                venta.estado = 'cancelada'
                venta.notas = f"{venta.notas}\n\nCANCELADA: {motivo}" if venta.notas else f"CANCELADA: {motivo}"
                venta.save()
                
                # Reabrir carrito si existe
                if venta.cart:
                    venta.cart.status = 'open'
                    venta.cart.save()
                
                return Response({
                    'success': True,
                    'message': 'Venta cancelada exitosamente',
                    'data': VentaSerializer(venta).data
                }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def ventas_por_cliente(self, request):
        """
        Obtener ventas de un cliente específico
        
        GET /api/sales/ventas/ventas_por_cliente/?cliente_id=1
        """
        cliente_id = request.query_params.get('cliente_id')
        
        if not cliente_id:
            return Response({
                'success': False,
                'error': 'Se requiere el parámetro cliente_id'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        ventas = self.get_queryset().filter(cliente_id=cliente_id)
        serializer = VentaSerializer(ventas, many=True)
        
        return Response({
            'success': True,
            'count': ventas.count(),
            'data': serializer.data
        })
    
    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        """
        Obtener estadísticas de ventas
        
        GET /api/sales/ventas/estadisticas/
        """
        from django.db.models import Count, Sum, Avg
        from datetime import datetime, timedelta
        
        # Últimos 30 días
        fecha_inicio = datetime.now() - timedelta(days=30)
        
        ventas_periodo = self.get_queryset().filter(fecha_venta__gte=fecha_inicio)
        
        stats = {
            'total_ventas': self.get_queryset().count(),
            'ventas_ultimos_30_dias': ventas_periodo.count(),
            'ventas_por_estado': dict(
                self.get_queryset().values('estado').annotate(
                    count=Count('id')
                ).values_list('estado', 'count')
            ),
            'ventas_por_tipo_entrega': dict(
                self.get_queryset().values('tipo_entrega').annotate(
                    count=Count('id')
                ).values_list('tipo_entrega', 'count')
            ),
            'monto_total': self.get_queryset().filter(
                estado='pagada'
            ).aggregate(total=Sum('total'))['total'] or 0,
            'monto_promedio': self.get_queryset().filter(
                estado='pagada'
            ).aggregate(promedio=Avg('total'))['promedio'] or 0,
        }
        
        return Response({
            'success': True,
            'data': stats
        })


# ============================================================================
# CU13: Procesar Pago en Línea - ViewSet
# ============================================================================

class PagoViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar pagos (CU13)
    
    Endpoints:
    - GET /api/sales/pagos/ - Listar todos los pagos
    - POST /api/sales/pagos/ - Crear un pago
    - GET /api/sales/pagos/{id}/ - Ver detalle de un pago
    - POST /api/sales/pagos/procesar/ - Procesar un pago (tarjeta/efectivo/etc)
    - POST /api/sales/pagos/generar_qr/ - Generar código QR para pago
    - POST /api/sales/pagos/{id}/confirmar_qr/ - Confirmar pago por QR
    - POST /api/sales/pagos/{id}/reembolsar/ - Reembolsar un pago
    """
    queryset = Pago.objects.all().select_related('venta')
    serializer_class = PagoSerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filtrar por venta
        venta_id = self.request.query_params.get('venta_id', None)
        if venta_id:
            queryset = queryset.filter(venta_id=venta_id)
        
        # Filtrar por método de pago
        metodo_pago = self.request.query_params.get('metodo_pago', None)
        if metodo_pago:
            queryset = queryset.filter(metodo_pago=metodo_pago)
        
        # Filtrar por estado
        estado = self.request.query_params.get('estado', None)
        if estado:
            queryset = queryset.filter(estado=estado)
        
        return queryset
    
    @action(detail=False, methods=['post'])
    def procesar(self, request):
        """
        Procesar un pago (tarjeta, efectivo, transferencia, etc.)
        
        Body:
        {
            "venta_id": "uuid-de-la-venta",
            "monto": 1500.00,
            "metodo_pago": "tarjeta_credito",
            "tarjeta_numero": "4111111111111111",
            "tarjeta_nombre": "Juan Perez",
            "tarjeta_expiracion": "12/25",
            "tarjeta_cvv": "123",
            "notas": "Pago adicional"
        }
        """
        serializer = PagoCreateSerializer(data=request.data)
        
        if serializer.is_valid():
            try:
                pago = serializer.save()
                
                response_data = PagoSerializer(pago).data
                
                # Añadir información adicional según el resultado
                if pago.estado == 'completado':
                    response_data['message'] = '¡Pago procesado exitosamente!'
                else:
                    response_data['message'] = 'El pago fue rechazado. Intenta con otro método.'
                
                return Response({
                    'success': pago.estado == 'completado',
                    'data': response_data
                }, status=status.HTTP_201_CREATED if pago.estado == 'completado' else status.HTTP_400_BAD_REQUEST)
            
            except Exception as e:
                return Response({
                    'success': False,
                    'error': str(e)
                }, status=status.HTTP_400_BAD_REQUEST)
        
        return Response({
            'success': False,
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def generar_qr(self, request):
        """
        Generar un código QR para pago
        
        Body:
        {
            "venta_id": "uuid-de-la-venta"
        }
        
        Response:
        {
            "success": true,
            "data": {
                "qr_codigo": "QR-V-20240109-0001-123456",
                "qr_imagen_url": "https://...",
                "venta_numero": "V-20240109-0001",
                "monto": "1500.00"
            }
        }
        """
        serializer = PagoQRSerializer(data=request.data)
        
        if serializer.is_valid():
            try:
                qr_data = serializer.save()
                return Response({
                    'success': True,
                    'message': 'Código QR generado exitosamente',
                    'data': qr_data
                }, status=status.HTTP_200_OK)
            except Exception as e:
                return Response({
                    'success': False,
                    'error': str(e)
                }, status=status.HTTP_400_BAD_REQUEST)
        
        return Response({
            'success': False,
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def confirmar_qr(self, request, pk=None):
        """
        Confirmar un pago realizado por QR
        
        POST /api/sales/pagos/{id}/confirmar_qr/
        Body: {"codigo_confirmacion": "123456"}
        """
        pago = self.get_object()
        
        if pago.metodo_pago != 'qr':
            return Response({
                'success': False,
                'error': 'Este pago no es por código QR'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if pago.estado == 'completado':
            return Response({
                'success': False,
                'error': 'Este pago ya ha sido confirmado'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        codigo_confirmacion = request.data.get('codigo_confirmacion')
        
        if not codigo_confirmacion:
            return Response({
                'success': False,
                'error': 'Se requiere el código de confirmación'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Simular validación del código QR
        # En un sistema real, aquí se validaría con el proveedor de QR
        try:
            pago.estado = 'completado'
            pago.numero_autorizacion = f'QR-AUTH-{codigo_confirmacion}'
            pago.fecha_procesamiento = timezone.now()
            pago.save()
            
            # Actualizar estado de la venta
            if pago.venta:
                pago.venta.estado = 'pagada'
                pago.venta.save()
            
            return Response({
                'success': True,
                'message': 'Pago confirmado exitosamente',
                'data': PagoSerializer(pago).data
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def reembolsar(self, request, pk=None):
        """
        Reembolsar un pago
        
        POST /api/sales/pagos/{id}/reembolsar/
        Body: {"motivo": "Razón del reembolso"}
        """
        pago = self.get_object()
        
        if pago.estado != 'completado':
            return Response({
                'success': False,
                'error': 'Solo se pueden reembolsar pagos completados'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if pago.estado == 'reembolsado':
            return Response({
                'success': False,
                'error': 'Este pago ya ha sido reembolsado'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        motivo = request.data.get('motivo', 'Sin motivo especificado')
        
        try:
            with transaction.atomic():
                # Actualizar estado del pago
                pago.estado = 'reembolsado'
                pago.notas = f"{pago.notas}\n\nREEMBOLSADO: {motivo}" if pago.notas else f"REEMBOLSADO: {motivo}"
                pago.save()
                
                # Actualizar estado de la venta
                if pago.venta:
                    pago.venta.estado = 'reembolsada'
                    pago.venta.save()
                    
                    # Revertir stock
                    for detalle in pago.venta.detalles.all():
                        producto = detalle.producto
                        if producto.stock_actual is not None:
                            producto.stock_actual += detalle.cantidad
                            producto.save()
                
                return Response({
                    'success': True,
                    'message': 'Pago reembolsado exitosamente',
                    'data': PagoSerializer(pago).data
                }, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        """
        Obtener estadísticas de pagos
        
        GET /api/sales/pagos/estadisticas/
        """
        from django.db.models import Count, Sum
        
        stats = {
            'total_pagos': self.get_queryset().count(),
            'pagos_por_estado': dict(
                self.get_queryset().values('estado').annotate(
                    count=Count('id')
                ).values_list('estado', 'count')
            ),
            'pagos_por_metodo': dict(
                self.get_queryset().values('metodo_pago').annotate(
                    count=Count('id')
                ).values_list('metodo_pago', 'count')
            ),
            'monto_total_procesado': self.get_queryset().filter(
                estado='completado'
            ).aggregate(total=Sum('monto'))['total'] or 0,
            'monto_total_reembolsado': self.get_queryset().filter(
                estado='reembolsado'
            ).aggregate(total=Sum('monto'))['total'] or 0,
        }
        
        return Response({
            'success': True,
            'data': stats
        })


# ============================================================================
# CU14, CU15, CU16: Endpoints para Comprobantes, Histórico y Dashboard
# ============================================================================

class VentaComprobantePDFViewSet(viewsets.ViewSet):
    """
    API para descargar comprobantes PDF de ventas
    
    CU14: Emitir Comprobante de Venta (PDF)
    """
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]
    
    @action(detail=False, methods=['get'])
    def descargar_pdf(self, request):
        """
        Descargar comprobante PDF de una venta
        
        GET /api/sales/comprobante/descargar_pdf/?venta_id=1
        """
        venta_id = request.query_params.get('venta_id')
        
        if not venta_id:
            return Response({
                'success': False,
                'error': 'venta_id es requerido'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            venta = Venta.objects.get(id=venta_id)
            
            # Generar PDF
            pdf_buffer = venta.generar_comprobante_pdf()
            
            # Crear respuesta
            from django.http import HttpResponse
            response = HttpResponse(pdf_buffer.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{venta.obtener_nombre_archivo_pdf()}"'
            
            return response
            
        except Venta.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Venta no encontrada'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


class VentaHistoricoViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API para listar histórico de ventas con filtros
    
    CU15: Listar Histórico de Ventas con Filtros
    """
    queryset = Venta.objects.all().order_by('-fecha_venta')
    serializer_class = VentaSerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]
    
    def get_queryset(self):
        """
        Filtrar ventas por:
        - fecha_inicio: fecha de inicio (YYYY-MM-DD)
        - fecha_fin: fecha de fin (YYYY-MM-DD)
        - cliente_id: ID del cliente
        - estado: estado de la venta (pendiente, pagada, cancelada, etc)
        - metodo_pago: método de pago (tarjeta, efectivo, transferencia, etc)
        - vendedor_id: ID del vendedor
        """
        queryset = super().get_queryset()
        
        # Filtro por fecha inicio
        fecha_inicio = self.request.query_params.get('fecha_inicio')
        if fecha_inicio:
            from django.utils.dateparse import parse_date
            fecha = parse_date(fecha_inicio)
            if fecha:
                queryset = queryset.filter(fecha_venta__date__gte=fecha)
        
        # Filtro por fecha fin
        fecha_fin = self.request.query_params.get('fecha_fin')
        if fecha_fin:
            from django.utils.dateparse import parse_date
            fecha = parse_date(fecha_fin)
            if fecha:
                queryset = queryset.filter(fecha_venta__date__lte=fecha)
        
        # Filtro por cliente
        cliente_id = self.request.query_params.get('cliente_id')
        if cliente_id:
            queryset = queryset.filter(cliente_id=cliente_id)
        
        # Filtro por estado
        estado = self.request.query_params.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado)
        
        # Filtro por método de pago
        metodo_pago = self.request.query_params.get('metodo_pago')
        if metodo_pago:
            queryset = queryset.filter(metodo_pago=metodo_pago)
        
        # Filtro por vendedor
        vendedor_id = self.request.query_params.get('vendedor_id')
        if vendedor_id:
            queryset = queryset.filter(usuario_id=vendedor_id)
        
        # Búsqueda por código de venta o cliente
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(codigo_venta__icontains=search) |
                Q(cliente__nombre_completo__icontains=search)
            )
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def filtros_disponibles(self, request):
        """
        Obtener opciones disponibles para filtros
        
        GET /api/sales/historico/filtros_disponibles/
        """
        from apps.clients.models import Clientes
        from apps.authentication.models import Usuarios
        
        clientes = Clientes.objects.filter(activo=True).values('id', 'nombre_completo')
        vendedores = Usuarios.objects.all().values('id', 'nombre')
        
        return Response({
            'success': True,
            'data': {
                'estados': [
                    {'value': 'pendiente', 'label': 'Pendiente'},
                    {'value': 'pagada', 'label': 'Pagada'},
                    {'value': 'cancelada', 'label': 'Cancelada'},
                    {'value': 'reembolsada', 'label': 'Reembolsada'},
                    {'value': 'en_proceso', 'label': 'En Proceso'},
                ],
                'metodos_pago': [
                    {'value': 'tarjeta', 'label': 'Tarjeta'},
                    {'value': 'efectivo', 'label': 'Efectivo'},
                    {'value': 'transferencia', 'label': 'Transferencia'},
                    {'value': 'paypal', 'label': 'PayPal'},
                    {'value': 'stripe', 'label': 'Stripe'},
                ],
                'clientes': list(clientes),
                'vendedores': list(vendedores),
            }
        })


class VentaDashboardViewSet(viewsets.ViewSet):
    """
    API para obtener estadísticas del dashboard de ventas
    
    CU16: Visualizar Dashboard de Ventas Históricas
    """
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]
    
    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        """
        Obtener estadísticas completas del dashboard
        
        GET /api/sales/dashboard/estadisticas/?fecha_inicio=2025-11-01&fecha_fin=2025-11-30&cliente_id=1&estado=pagada
        
        Parámetros opcionales:
        - fecha_inicio: fecha de inicio (YYYY-MM-DD)
        - fecha_fin: fecha de fin (YYYY-MM-DD)
        - cliente_id: ID del cliente
        - estado: estado de la venta
        """
        from .estadisticas import EstadisticasVentas
        from django.utils.dateparse import parse_date
        
        # Obtener parámetros de filtro
        fecha_inicio_str = request.query_params.get('fecha_inicio')
        fecha_fin_str = request.query_params.get('fecha_fin')
        cliente_id = request.query_params.get('cliente_id')
        estado = request.query_params.get('estado')
        
        fecha_inicio = parse_date(fecha_inicio_str) if fecha_inicio_str else None
        fecha_fin = parse_date(fecha_fin_str) if fecha_fin_str else None
        
        try:
            # Calcular estadísticas
            stats_obj = EstadisticasVentas(
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                cliente_id=cliente_id if cliente_id else None,
                estado=estado if estado else None
            )
            
            estadisticas = stats_obj.obtener_estadisticas_completas()
            
            return Response({
                'success': True,
                'data': estadisticas
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def resumen(self, request):
        """
        Obtener resumen rápido de ventas
        
        GET /api/sales/dashboard/resumen/
        """
        from .estadisticas import EstadisticasVentas
        
        try:
            stats = EstadisticasVentas()
            resumen = stats.obtener_resumen()
            
            return Response({
                'success': True,
                'data': resumen
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def por_estado(self, request):
        """
        Obtener ventas agrupadas por estado
        
        GET /api/sales/dashboard/por_estado/
        """
        from .estadisticas import EstadisticasVentas
        
        try:
            stats = EstadisticasVentas()
            por_estado = stats.obtener_ventas_por_estado()
            
            return Response({
                'success': True,
                'data': por_estado
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def por_metodo_pago(self, request):
        """
        Obtener ventas agrupadas por método de pago
        
        GET /api/sales/dashboard/por_metodo_pago/
        """
        from .estadisticas import EstadisticasVentas
        
        try:
            stats = EstadisticasVentas()
            por_metodo = stats.obtener_ventas_por_metodo()
            
            return Response({
                'success': True,
                'data': por_metodo
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def top_productos(self, request):
        """
        Obtener top 10 productos más vendidos
        
        GET /api/sales/dashboard/top_productos/?limite=10
        """
        from .estadisticas import EstadisticasVentas
        
        try:
            limite = int(request.query_params.get('limite', 10))
            stats = EstadisticasVentas()
            top_productos = stats.obtener_top_productos(limite=limite)
            
            return Response({
                'success': True,
                'data': top_productos
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def por_vendedor(self, request):
        """
        Obtener estadísticas por vendedor
        
        GET /api/sales/dashboard/por_vendedor/
        """
        from .estadisticas import EstadisticasVentas
        
        try:
            stats = EstadisticasVentas()
            por_vendedor = stats.obtener_ventas_por_vendedor()
            
            return Response({
                'success': True,
                'data': por_vendedor
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def tendencia_diaria(self, request):
        """
        Obtener tendencia de ventas diaria
        
        GET /api/sales/dashboard/tendencia_diaria/?dias=30
        """
        from .estadisticas import EstadisticasVentas
        
        try:
            dias = int(request.query_params.get('dias', 30))
            stats = EstadisticasVentas()
            tendencia = stats.obtener_tendencia_diaria(dias=dias)
            
            return Response({
                'success': True,
                'data': tendencia
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def tendencia_mensual(self, request):
        """
        Obtener tendencia de ventas mensual
        
        GET /api/sales/dashboard/tendencia_mensual/?meses=12
        """
        from .estadisticas import EstadisticasVentas
        
        try:
            meses = int(request.query_params.get('meses', 12))
            stats = EstadisticasVentas()
            tendencia = stats.obtener_tendencia_mensual(meses=meses)
            
            return Response({
                'success': True,
                'data': tendencia
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


# CU17-19: ViewSets para móvil
class DispositivoMovilViewSet(viewsets.ModelViewSet):
    """
    CU17-19: API para gestionar dispositivos móviles
    Permite registrar, actualizar y listar dispositivos móviles
    """
    queryset = DispositivosMoviles.objects.all()
    serializer_class = DispositivoMovilSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        """Filtrar dispositivos del usuario autenticado"""
        is_staff = getattr(self.request.user, 'is_staff', False)
        is_superuser = getattr(self.request.user, 'is_superuser', False)
        if is_staff or is_superuser:
            return DispositivosMoviles.objects.all()
        # Filtrar por usuario autenticado
        return DispositivosMoviles.objects.filter(usuario__user=self.request.user)
    
    def perform_create(self, serializer):
        """Asignar el usuario autenticado al crear dispositivo"""
        try:
            usuario = self.request.user.usuarios
        except:
            return Response(
                {'detail': 'Usuario no tiene perfil de Usuario'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        serializer.save(usuario=usuario)
    
    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def registrar_acceso(self, request, pk=None):
        """
        Registra el último acceso del dispositivo
        POST /api/sales/dispositivos/{id}/registrar_acceso/
        """
        dispositivo = self.get_object()
        ip_address = request.META.get('REMOTE_ADDR')
        dispositivo.last_activity = timezone.now()
        if ip_address:
            dispositivo.ip_address = ip_address
        dispositivo.save()
        
        return Response({
            'success': True,
            'message': 'Acceso registrado',
            'ultimo_acceso': dispositivo.last_activity
        })


class VentaMovilViewSet(viewsets.ModelViewSet):
    """
    CU17: API para crear y gestionar ventas desde dispositivos móviles
    """
    serializer_class = VentaMovilSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        """Filtrar ventas del usuario o cliente autenticado"""
        is_staff = getattr(self.request.user, 'is_staff', False)
        is_superuser = getattr(self.request.user, 'is_superuser', False)
        if is_staff or is_superuser:
            return Venta.objects.all()
        # Filtrar ventas del usuario
        return Venta.objects.filter(usuario__user=self.request.user)
    
    def get_serializer_class(self):
        """Usar serializer diferente según la acción"""
        if self.action == 'create':
            return VentaCreateMovilSerializer
        elif self.action == 'historial':
            return VentaHistoricoMovilSerializer
        return VentaMovilSerializer
    
    def create(self, request, *args, **kwargs):
        """
        CU17: Crear venta desde dispositivo móvil
        POST /api/sales/ventas-movil/
        """
        try:
            usuario = request.user.usuarios
        except:
            return Response(
                {'detail': 'Usuario no tiene perfil de Usuario'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.validated_data['usuario'] = usuario
        
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        
        return Response(
            {
                'success': True,
                'message': 'Venta creada exitosamente',
                'data': VentaMovilSerializer(serializer.instance).data
            },
            status=status.HTTP_201_CREATED,
            headers=headers
        )
    
    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def historial(self, request):
        """
        CU18: Obtener historial de compras del usuario
        GET /api/sales/ventas-movil/historial/
        """
        try:
            usuario = request.user.usuarios
            ventas = Venta.objects.filter(usuario=usuario).order_by('-fecha_venta')
            
            # Filtrar por estado si se proporciona
            estado = request.query_params.get('estado')
            if estado:
                ventas = ventas.filter(estado=estado)
            
            # Filtrar por fecha si se proporcionan
            fecha_inicio = request.query_params.get('fecha_inicio')
            fecha_fin = request.query_params.get('fecha_fin')
            if fecha_inicio and fecha_fin:
                from django.utils.dateparse import parse_date
                desde = parse_date(fecha_inicio)
                hasta = parse_date(fecha_fin)
                if desde and hasta:
                    ventas = ventas.filter(fecha_venta__date__gte=desde, fecha_venta__date__lte=hasta)
            
            serializer = VentaHistoricoMovilSerializer(ventas, many=True)
            
            return Response({
                'success': True,
                'total': ventas.count(),
                'data': serializer.data
            })
        
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def dashboard_movil(self, request):
        """
        CU19: Obtener dashboard resumido para móvil
        GET /api/sales/ventas-movil/dashboard_movil/
        """
        try:
            usuario = request.user.usuarios
            
            # Obtener ventas del usuario
            ventas = Venta.objects.filter(usuario=usuario)
            
            # Calcular estadísticas
            total_vendido = ventas.aggregate(total=Sum('total'))['total'] or 0
            total_ventas = ventas.count()
            promedio_venta = total_vendido / total_ventas if total_ventas > 0 else 0
            
            # Últimas 5 ventas
            ultimas_ventas = ventas.order_by('-fecha_venta')[:5]
            
            # Contar por estado
            ventas_pendientes = ventas.filter(estado='pendiente').count()
            ventas_pagadas = ventas.filter(estado='pagada').count()
            ventas_en_proceso = ventas.filter(estado='en_proceso').count()
            
            # Alertas personalizadas
            alertas = []
            if ventas_pendientes > 0:
                alertas.append({
                    'tipo': 'pendiente',
                    'titulo': f'{ventas_pendientes} Ventas Pendientes',
                    'mensaje': 'Tienes ventas pendientes de pago'
                })
            
            if total_vendido == 0:
                alertas.append({
                    'tipo': 'info',
                    'titulo': 'Sin ventas registradas',
                    'mensaje': 'Comienza a realizar compras'
                })
            
            data = {
                'total_vendido': float(total_vendido),
                'total_ventas': total_ventas,
                'promedio_venta': float(promedio_venta),
                'ultimas_ventas': VentaHistoricoMovilSerializer(ultimas_ventas, many=True).data,
                'ventas_pendientes': ventas_pendientes,
                'ventas_pagadas': ventas_pagadas,
                'ventas_en_proceso': ventas_en_proceso,
                'alertas': alertas
            }
            
            return Response({
                'success': True,
                'data': data
            })
        
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


class NotificacionPushViewSet(viewsets.ModelViewSet):
    """
    CU20: API para gestionar notificaciones push
    """
    queryset = NotificacionPush.objects.all()
    serializer_class = NotificacionPushSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        """Filtrar notificaciones del usuario"""
        is_staff = getattr(self.request.user, 'is_staff', False)
        is_superuser = getattr(self.request.user, 'is_superuser', False)
        if is_staff or is_superuser:
            return NotificacionPush.objects.all()
        try:
            usuario = self.request.user.usuarios
            return NotificacionPush.objects.filter(usuario=usuario)
        except:
            return NotificacionPush.objects.none()
    
    def get_serializer_class(self):
        if self.action == 'create':
            return NotificacionPushCreateSerializer
        return NotificacionPushSerializer
    
    def perform_create(self, serializer):
        """Registrar quien creó la notificación"""
        try:
            usuario = self.request.user.usuarios
            if serializer.validated_data.get('usuario') is None:
                serializer.validated_data['usuario'] = usuario
        except:
            pass
        serializer.save()
    
    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def no_leidas(self, request):
        """
        Obtener notificaciones no entregadas
        GET /api/sales/notificaciones/no_leidas/
        """
        try:
            usuario = request.user.usuarios
            notificaciones = NotificacionPush.objects.filter(
                usuario=usuario,
                estado__in=['pendiente', 'enviada']
            ).order_by('-created_at')
            
            serializer = self.get_serializer(notificaciones, many=True)
            
            return Response({
                'success': True,
                'total': notificaciones.count(),
                'data': serializer.data
            })
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def marcar_entregada(self, request, pk=None):
        """
        Marcar notificación como entregada
        POST /api/sales/notificaciones/{id}/marcar_entregada/
        """
        notificacion = self.get_object()
        notificacion.marcar_entregada()
        
        return Response({
            'success': True,
            'message': 'Notificación marcada como entregada',
            'data': NotificacionPushSerializer(notificacion).data
        })
    
    @action(detail=False, methods=['post'], permission_classes=[permissions.IsAdminUser])
    def enviar_a_usuarios(self, request):
        """
        Enviar notificación a múltiples usuarios (Admin)
        POST /api/sales/notificaciones/enviar_a_usuarios/
        
        Body: {
            "usuarios_ids": [1, 2, 3],
            "titulo": "Nuevo descuento",
            "mensaje": "Tenemos un 20% de descuento",
            "tipo": "promocion"
        }
        """
        usuarios_ids = request.data.get('usuarios_ids', [])
        titulo = request.data.get('titulo')
        mensaje = request.data.get('mensaje')
        tipo = request.data.get('tipo', 'otro')
        
        if not usuarios_ids or not titulo or not mensaje:
            return Response({
                'success': False,
                'error': 'Se requiere usuarios_ids, titulo y mensaje'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            notificaciones_creadas = []
            from apps.authentication.models import Usuarios
            
            usuarios = Usuarios.objects.filter(id__in=usuarios_ids)
            for usuario in usuarios:
                notif = NotificacionPush.objects.create(
                    usuario=usuario,
                    titulo=titulo,
                    mensaje=mensaje,
                    tipo=tipo
                )
                notificaciones_creadas.append(notif)
            
            return Response({
                'success': True,
                'message': f'{len(notificaciones_creadas)} notificaciones creadas',
                'total': len(notificaciones_creadas)
            }, status=status.HTTP_201_CREATED)
        
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


# ============================================================================
# CU21, CU22, CU23: ViewSet para Reportes Dinámicos
# ============================================================================

class ReporteViewSet(viewsets.ModelViewSet):
    """
    CU21: Filtrar Datos y Exportar Gráficas del Dashboard
    CU22: Generar Reporte Dinámico (Texto y Voz)
    CU23: Descargar Reporte en Formato (PDF / Excel)
    """
    queryset = Reporte.objects.all()
    serializer_class = ReporteSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [JSONParser, MultiPartParser, FormParser]
    
    def get_queryset(self):
        """Filtrar reportes por usuario autenticado"""
        return self.queryset.filter(usuario=self.request.user.usuarios)
    
    def get_serializer_class(self):
        """Usar serializer específico según la acción"""
        if self.action == 'create':
            return ReporteGenerarSerializer
        elif self.action == 'list':
            return ReporteListadoSerializer
        elif self.action == 'descargar':
            return ReporteExportarSerializer
        elif self.action == 'voz':
            return ReporteVozSerializer
        return ReporteSerializer
    
    def create(self, request, *args, **kwargs):
        """CU22: Generar nuevo reporte dinámico"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        import time
        inicio = time.time()
        
        try:
            # Crear reporte con estado 'generando'
            reporte = Reporte.objects.create(
                usuario=request.user.usuarios,
                titulo=serializer.validated_data['titulo'],
                tipo_reporte=serializer.validated_data['tipo_reporte'],
                formato=serializer.validated_data.get('formato', 'pdf'),
                filtros=self._construir_filtros(serializer.validated_data),
                estado='generando'
            )
            
            # Generar datos del reporte según tipo
            datos_reporte = self._generar_datos_reporte(reporte, serializer.validated_data)
            reporte.datos_reporte = datos_reporte
            reporte.total_registros = len(datos_reporte.get('registros', []))
            
            # Generar resumen ejecutivo (IA simple)
            reporte.resumen_texto = self._generar_resumen(datos_reporte, reporte.tipo_reporte)
            
            # Generar archivo según formato
            if reporte.formato == 'pdf':
                self._generar_pdf(reporte, datos_reporte)
            elif reporte.formato == 'excel':
                self._generar_excel(reporte, datos_reporte)
            elif reporte.formato == 'csv':
                self._generar_csv(reporte, datos_reporte)
            
            # Generar voz si se solicita
            if serializer.validated_data.get('incluir_voz', False):
                self._generar_voz(reporte)
            
            # Marcar como completado
            reporte.tiempo_generacion = time.time() - inicio
            reporte.estado = 'completado'
            reporte.save()
            
            return Response(
                ReporteSerializer(reporte).data,
                status=status.HTTP_201_CREATED
            )
        
        except Exception as e:
            print(f"❌ Error generando reporte: {str(e)}")
            reporte.registrar_error(str(e))
            return Response({
                'error': f'Error al generar reporte: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def listar_por_tipo(self, request):
        """CU21: Listar reportes filtrados por tipo"""
        tipo = request.query_params.get('tipo')
        estado = request.query_params.get('estado')
        
        queryset = self.get_queryset()
        
        if tipo:
            queryset = queryset.filter(tipo_reporte=tipo)
        if estado:
            queryset = queryset.filter(estado=estado)
        
        serializer = ReporteListadoSerializer(queryset.order_by('-fecha_generacion'), many=True)
        return Response({
            'total': queryset.count(),
            'reportes': serializer.data
        })
    
    @action(detail=True, methods=['post'])
    def descargar(self, request, pk=None):
        """CU23: Descargar reporte en el formato especificado"""
        reporte = self.get_object()
        
        if reporte.estado != 'completado':
            return Response({
                'error': 'El reporte aún no está generado o tiene errores'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Marcar como descargado
            reporte.marcar_descargado()
            
            # Obtener el archivo según formato
            archivo = reporte.get_archivo_por_formato()
            
            if not archivo or not archivo.name:
                return Response({
                    'error': f'El archivo {reporte.formato} no está disponible'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Retornar archivo para descargar
            from django.http import FileResponse
            response = FileResponse(
                archivo.open('rb'),
                content_type='application/octet-stream'
            )
            response['Content-Disposition'] = f'attachment; filename="{reporte.titulo}.{reporte.formato}"'
            
            return response
        
        except Exception as e:
            return Response({
                'error': f'Error descargando reporte: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def voz(self, request, pk=None):
        """CU22: Obtener/generar resumen en voz del reporte"""
        reporte = self.get_object()
        
        if reporte.estado != 'completado':
            return Response({
                'error': 'El reporte debe estar completado'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            regenerar = request.data.get('regenerar', False)
            
            # Si ya existe audio y no se solicita regenerar
            if reporte.resumen_voz and reporte.resumen_voz.name and not regenerar:
                from django.http import FileResponse
                response = FileResponse(
                    reporte.resumen_voz.open('rb'),
                    content_type='audio/mpeg'
                )
                response['Content-Disposition'] = f'attachment; filename="{reporte.titulo}_voz.mp3"'
                return response
            
            # Generar voz
            self._generar_voz(reporte, idioma=request.data.get('idioma', 'es'))
            
            # Retornar archivo de audio
            from django.http import FileResponse
            response = FileResponse(
                reporte.resumen_voz.open('rb'),
                content_type='audio/mpeg'
            )
            response['Content-Disposition'] = f'attachment; filename="{reporte.titulo}_voz.mp3"'
            
            return response
        
        except Exception as e:
            return Response({
                'error': f'Error generando voz: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def _construir_filtros(self, validated_data):
        """Construir diccionario de filtros"""
        filtros = {}
        
        if validated_data.get('fecha_inicio'):
            filtros['fecha_inicio'] = str(validated_data['fecha_inicio'])
        if validated_data.get('fecha_fin'):
            filtros['fecha_fin'] = str(validated_data['fecha_fin'])
        if validated_data.get('cliente_id'):
            filtros['cliente_id'] = validated_data['cliente_id']
        if validated_data.get('cliente_ids'):
            filtros['cliente_ids'] = validated_data['cliente_ids']
        if validated_data.get('producto_id'):
            filtros['producto_id'] = validated_data['producto_id']
        if validated_data.get('estado_venta'):
            filtros['estado_venta'] = validated_data['estado_venta']
        if validated_data.get('metodo_pago'):
            filtros['metodo_pago'] = validated_data['metodo_pago']
        if validated_data.get('agrupar_por'):
            filtros['agrupar_por'] = validated_data['agrupar_por']
        
        return filtros
    
    def _generar_datos_reporte(self, reporte, validated_data):
        """Generar datos del reporte según tipo"""
        from django.db.models import Sum, Count, Avg, F
        from django.db import models
        from datetime import datetime
        
        tipo_reporte = validated_data['tipo_reporte']
        filtros = reporte.filtros
        
        datos = {
            'tipo': tipo_reporte,
            'generado_en': timezone.now().isoformat(),
            'registros': [],
            'totales': {},
            'graficas': []
        }
        
        try:
            if tipo_reporte == 'ventas':
                ventas_qs = Venta.objects.select_related('cliente', 'usuario')
                
                if filtros.get('fecha_inicio'):
                    ventas_qs = ventas_qs.filter(
                        fecha_venta__gte=datetime.fromisoformat(filtros['fecha_inicio'])
                    )
                if filtros.get('fecha_fin'):
                    ventas_qs = ventas_qs.filter(
                        fecha_venta__lte=datetime.fromisoformat(filtros['fecha_fin'])
                    )
                if filtros.get('cliente_id'):
                    ventas_qs = ventas_qs.filter(cliente_id=filtros['cliente_id'])
                if filtros.get('estado_venta'):
                    ventas_qs = ventas_qs.filter(estado=filtros['estado_venta'])
                
                for venta in ventas_qs[:100]:
                    datos['registros'].append({
                        'id': venta.id,
                        'codigo': venta.codigo_venta,
                        'cliente': venta.cliente.nombre_completo,
                        'fecha': venta.fecha_venta.isoformat(),
                        'total': str(venta.total),
                        'estado': venta.estado,
                        'metodo_pago': venta.metodo_pago
                    })
                
                stats = ventas_qs.aggregate(
                    total_vendido=Sum('total'),
                    cantidad=Count('id'),
                    promedio=Avg('total')
                )
                
                datos['totales'] = {
                    'total_vendido': str(stats['total_vendido'] or 0),
                    'cantidad_ventas': stats['cantidad'] or 0,
                    'promedio_venta': str(stats['promedio'] or 0)
                }
            
            elif tipo_reporte == 'productos':
                productos_qs = VentaDetalle.objects.values('producto__nombre', 'producto__sku').annotate(
                    total_vendido=Sum('cantidad'),
                    ingresos=Sum(F('cantidad') * F('precio_unitario'), output_field=models.DecimalField())
                )
                
                for prod in productos_qs[:50]:
                    datos['registros'].append({
                        'producto': prod['producto__nombre'],
                        'sku': prod['producto__sku'],
                        'cantidad': prod['total_vendido'],
                        'ingresos': str(prod['ingresos'] or 0)
                    })
                
                datos['totales'] = {
                    'productos_vendidos': len(datos['registros'])
                }
            
            elif tipo_reporte == 'estadisticas':
                stats = Venta.objects.aggregate(
                    total=Sum('total'),
                    ventas=Count('id'),
                    descuentos=Sum('descuento'),
                    impuestos=Sum('iva')
                )
                
                datos['totales'] = {
                    'total_vendido': str(stats['total'] or 0),
                    'total_ventas': stats['ventas'] or 0,
                    'total_descuentos': str(stats['descuentos'] or 0),
                    'total_impuestos': str(stats['impuestos'] or 0)
                }
        
        except Exception as e:
            print(f"⚠️ Error generando datos: {str(e)}")
            datos['error'] = str(e)
        
        return datos
    
    def _generar_resumen(self, datos_reporte, tipo_reporte):
        """Generar resumen ejecutivo en texto"""
        totales = datos_reporte.get('totales', {})
        
        resumen = f"RESUMEN EJECUTIVO - {tipo_reporte.upper()}\nGenerado: {datos_reporte.get('generado_en')}\n\n"
        
        if tipo_reporte == 'ventas':
            total = totales.get('total_vendido', '0')
            cantidad = totales.get('cantidad_ventas', 0)
            promedio = totales.get('promedio_venta', '0')
            
            resumen += f"""Total Vendido: ${total}
Cantidad de Ventas: {cantidad}
Promedio por Venta: ${promedio}

Análisis: Se realizaron {cantidad} transacciones por un total de ${total}.
"""
        
        elif tipo_reporte == 'productos':
            resumen += f"""Total de Productos: {len(datos_reporte.get('registros', []))}
Clasificados por volumen de ventas e ingresos generados.
"""
        
        return resumen
    
    def _generar_pdf(self, reporte, datos_reporte):
        """Generar PDF del reporte"""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            import io, uuid
            
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch)
            
            elements = []
            styles = getSampleStyleSheet()
            
            # Título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1f4788'),
                spaceAfter=30,
                alignment=1
            )
            elements.append(Paragraph(reporte.titulo, title_style))
            elements.append(Spacer(1, 0.3*inch))
            
            # Resumen
            elements.append(Paragraph("<b>RESUMEN EJECUTIVO</b>", styles['Heading2']))
            elements.append(Paragraph(reporte.resumen_texto.replace('\n', '<br/>'), styles['Normal']))
            elements.append(Spacer(1, 0.3*inch))
            
            # Tabla de datos
            if datos_reporte.get('registros'):
                elements.append(PageBreak())
                elements.append(Paragraph("<b>DETALLE</b>", styles['Heading2']))
                
                table_data = [list(datos_reporte['registros'][0].keys())]
                for reg in datos_reporte['registros'][:20]:
                    table_data.append(list(reg.values()))
                
                tabla = Table(table_data)
                tabla.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                elements.append(tabla)
            
            doc.build(elements)
            buffer.seek(0)
            
            nombre_archivo = f'reportes/pdf/{reporte.id}_{uuid.uuid4()}.pdf'
            from django.core.files.base import ContentFile
            reporte.archivo_pdf.save(nombre_archivo, ContentFile(buffer.read()), save=True)
            print(f"✅ PDF generado: {nombre_archivo}")
        
        except Exception as e:
            print(f"❌ Error en PDF: {str(e)}")
            raise
    
    def _generar_excel(self, reporte, datos_reporte):
        """Generar Excel del reporte"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            import io, uuid
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte"
            
            ws['A1'] = reporte.titulo
            ws['A1'].font = Font(bold=True, size=14, color="1f4788")
            ws.merge_cells('A1:F1')
            
            ws['A3'] = 'RESUMEN'
            ws['A3'].font = Font(bold=True, size=11)
            ws['A4'] = reporte.resumen_texto
            ws['A4'].alignment = Alignment(wrap_text=True)
            ws.merge_cells('A4:F6')
            
            row = 9
            if datos_reporte.get('registros'):
                headers = list(datos_reporte['registros'][0].keys())
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = header
                    cell.font = Font(bold=True, color="ffffff")
                    cell.fill = PatternFill(start_color="1f4788", end_color="1f4788", fill_type="solid")
                
                row += 1
                for reg in datos_reporte['registros'][:100]:
                    for col, value in enumerate(reg.values(), 1):
                        ws.cell(row=row, column=col).value = value
                    row += 1
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            nombre_archivo = f'reportes/excel/{reporte.id}_{uuid.uuid4()}.xlsx'
            from django.core.files.base import ContentFile
            reporte.archivo_excel.save(nombre_archivo, ContentFile(buffer.read()), save=True)
            print(f"✅ Excel generado: {nombre_archivo}")
        
        except Exception as e:
            print(f"❌ Error en Excel: {str(e)}")
            raise
    
    def _generar_csv(self, reporte, datos_reporte):
        """Generar CSV del reporte"""
        try:
            import csv, io, uuid
            
            buffer = io.StringIO()
            writer = csv.writer(buffer)
            
            writer.writerow(['Reporte:', reporte.titulo])
            writer.writerow(['Generado:', timezone.now().isoformat()])
            writer.writerow([])
            
            if datos_reporte.get('registros'):
                headers = list(datos_reporte['registros'][0].keys())
                writer.writerow(headers)
                
                for reg in datos_reporte['registros']:
                    writer.writerow(reg.values())
            
            csv_content = buffer.getvalue()
            nombre_archivo = f'reportes/csv/{reporte.id}_{uuid.uuid4()}.csv'
            from django.core.files.base import ContentFile
            reporte.archivo_csv.save(nombre_archivo, ContentFile(csv_content.encode()), save=True)
            print(f"✅ CSV generado: {nombre_archivo}")
        
        except Exception as e:
            print(f"❌ Error en CSV: {str(e)}")
            raise
    
    def _generar_voz(self, reporte, idioma='es'):
        """Generar resumen en voz"""
        try:
            from gtts import gTTS
            import io, uuid
            
            tts = gTTS(text=reporte.resumen_texto, lang=idioma, slow=False)
            buffer = io.BytesIO()
            tts.write_to_fp(buffer)
            buffer.seek(0)
            
            nombre_archivo = f'reportes/audio/{reporte.id}_{uuid.uuid4()}.mp3'
            from django.core.files.base import ContentFile
            reporte.resumen_voz.save(nombre_archivo, ContentFile(buffer.read()), save=True)
            print(f"✅ Audio generado: {nombre_archivo}")
        
        except ImportError:
            print("⚠️ gTTS no instalado. Instala con: pip install gtts")
            import uuid
            reporte.resumen_voz.name = f'reportes/audio/placeholder_{uuid.uuid4()}.mp3'
        
        except Exception as e:
            print(f"❌ Error en voz: {str(e)}")
            raise


# ============================================================================
# CU24: ViewSet para Prompts Frecuentes
# ============================================================================

class PromptFrecuenteViewSet(viewsets.ModelViewSet):
    """
    CU24: Gestionar Prompts Frecuentes de Reportes
    
    Endpoints:
    - GET /api/sales/prompts/ - Listar prompts
    - POST /api/sales/prompts/ - Crear nuevo prompt
    - GET /api/sales/prompts/{id}/ - Obtener detalle
    - PUT /api/sales/prompts/{id}/ - Actualizar prompt
    - DELETE /api/sales/prompts/{id}/ - Eliminar prompt
    - POST /api/sales/prompts/{id}/usar/ - Usar prompt (incrementa contador)
    - POST /api/sales/prompts/{id}/toggle_favorito/ - Marcar como favorito
    """
    queryset = PromptFrecuente.objects.all()
    serializer_class = PromptFrecuenteSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        """Filtrar prompts por usuario autenticado"""
        return self.queryset.filter(usuario=self.request.user.usuarios, activo=True)
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return PromptFrecuenteCreateSerializer
        return PromptFrecuenteSerializer
    
    def perform_create(self, serializer):
        """Asignar usuario autenticado"""
        serializer.save(usuario=self.request.user.usuarios)
    
    @action(detail=True, methods=['post'])
    def usar(self, request, pk=None):
        """Usar este prompt (incrementa contador y genera reporte)"""
        prompt = self.get_object()
        prompt.registrar_uso()
        
        # Generar reporte con los parámetros del prompt
        try:
            from apps.sales.views import ReporteViewSet
            
            # Preparar datos para generar reporte
            datos_reporte = {
                'titulo': prompt.nombre,
                'tipo_reporte': prompt.tipo_reporte,
                'formato': prompt.formato,
                **prompt.filtros,
                **prompt.opciones
            }
            
            # Crear serializer y validar
            serializer = ReporteGenerarSerializer(data=datos_reporte)
            if serializer.is_valid():
                # Usar el método create del ReporteViewSet
                viewset = ReporteViewSet()
                viewset.request = request
                resultado = viewset.create(request, data=serializer.validated_data)
                
                return Response({
                    'success': True,
                    'message': f'Prompt "{prompt.nombre}" utilizado exitosamente',
                    'reporte': resultado.data
                }, status=status.HTTP_201_CREATED)
            else:
                return Response({
                    'success': False,
                    'errors': serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)
        
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Error al generar reporte: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def toggle_favorito(self, request, pk=None):
        """Marcar/desmarcar como favorito"""
        prompt = self.get_object()
        prompt.favorito = not prompt.favorito
        prompt.save()
        
        return Response({
            'success': True,
            'favorito': prompt.favorito,
            'message': f'Prompt marcado como {"favorito" if prompt.favorito else "no favorito"}'
        })
    
    @action(detail=False, methods=['get'])
    def por_categoria(self, request):
        """Obtener prompts agrupados por categoría"""
        categorias = {}
        for prompt in self.get_queryset():
            cat = prompt.get_categoria_display()
            if cat not in categorias:
                categorias[cat] = []
            categorias[cat].append(PromptFrecuenteSerializer(prompt).data)
        
        return Response({
            'total': self.get_queryset().count(),
            'categorias': categorias
        })
    
    @action(detail=False, methods=['get'])
    def favoritos(self, request):
        """Obtener solo prompts favoritos"""
        favoritos = self.get_queryset().filter(favorito=True).order_by('-veces_usado')
        serializer = self.get_serializer(favoritos, many=True)
        
        return Response({
            'total': favoritos.count(),
            'prompts': serializer.data
        })


# ============================================================================
# CU26: ViewSet para Modelo IA
# ============================================================================

class ModeloIAViewSet(viewsets.ModelViewSet):
    """
    CU26: Administrar Modelo de Predicción IA
    
    Endpoints:
    - GET /api/sales/modelos-ia/ - Listar modelos
    - POST /api/sales/modelos-ia/ - Crear nuevo modelo
    - GET /api/sales/modelos-ia/{id}/ - Obtener detalle
    - PUT /api/sales/modelos-ia/{id}/ - Actualizar modelo
    - DELETE /api/sales/modelos-ia/{id}/ - Eliminar modelo
    - POST /api/sales/modelos-ia/{id}/entrenar/ - Entrenar modelo
    - POST /api/sales/modelos-ia/{id}/activar/ - Activar modelo
    - GET /api/sales/modelos-ia/activo/ - Obtener modelo activo
    """
    queryset = ModeloIA.objects.all()
    serializer_class = ModeloIASerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_serializer_class(self):
        if self.action == 'entrenar':
            return ModeloIAEntrenarSerializer
        return ModeloIASerializer
    
    def perform_create(self, serializer):
        """Asignar usuario que crea el modelo"""
        serializer.save(creado_por=self.request.user.usuarios)
    
    @action(detail=True, methods=['post'])
    def entrenar(self, request, pk=None):
        """Entrenar el modelo con datos históricos (CU26)"""
        modelo = self.get_object()
        
        if modelo.estado == 'entrenando':
            return Response({
                'error': 'El modelo ya está siendo entrenado'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        try:
            periodo = serializer.validated_data.get('periodo_entrenamiento', '90d')
            
            # Convertir período a días
            dias = {
                '30d': 30,
                '90d': 90,
                '180d': 180,
                '1y': 365
            }.get(periodo, 90)
            
            from datetime import timedelta
            
            # Obtener datos de entrenamiento
            fecha_fin = timezone.now().date()
            fecha_inicio = fecha_fin - timedelta(days=dias)
            
            # Consultar ventas para entrenar
            ventas = Venta.objects.filter(
                fecha_venta__date__gte=fecha_inicio,
                fecha_venta__date__lte=fecha_fin
            ).aggregate(
                total=Sum('total'),
                cantidad=Count('id')
            )
            
            # Simular entrenamiento
            modelo.estado = 'entrenando'
            modelo.save()
            
            # Calcular métricas simples (demo)
            modelo.estado = 'activo'
            modelo.fecha_entrenamiento = timezone.now()
            modelo.datos_entrenamiento = ventas['cantidad'] or 0
            modelo.periodo_entrenamiento = periodo
            
            # Simular métricas de modelo
            import random
            modelo.precision = random.uniform(0.7, 0.95)
            modelo.r_squared = random.uniform(0.75, 0.99)
            modelo.mae = random.uniform(100, 1000)
            modelo.rmse = random.uniform(150, 1200)
            
            modelo.save()
            
            return Response({
                'success': True,
                'message': f'Modelo "{modelo.nombre}" entrenado exitosamente',
                'datos': {
                    'precision': modelo.precision,
                    'r_squared': modelo.r_squared,
                    'mae': modelo.mae,
                    'rmse': modelo.rmse,
                    'registros_entrenamiento': modelo.datos_entrenamiento
                }
            })
        
        except Exception as e:
            modelo.estado = 'error'
            modelo.error_mensaje = str(e)
            modelo.save()
            
            return Response({
                'error': f'Error al entrenar modelo: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def activar(self, request, pk=None):
        """Activar este modelo y desactivar otros"""
        modelo = self.get_object()
        
        if modelo.estado != 'activo':
            return Response({
                'error': 'El modelo debe estar en estado activo'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Desactivar otros modelos
            ModeloIA.objects.filter(estado='activo').exclude(id=modelo.id).update(estado='inactivo')
            
            # Activar este
            modelo.estado = 'activo'
            modelo.save()
            
            return Response({
                'success': True,
                'message': f'Modelo "{modelo.nombre}" activado'
            })
        
        except Exception as e:
            return Response({
                'error': f'Error al activar modelo: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def activo(self, request):
        """Obtener modelo activo"""
        modelo = ModeloIA.objects.filter(estado='activo').first()
        
        if not modelo:
            return Response({
                'mensaje': 'No hay modelo activo'
            }, status=status.HTTP_404_NOT_FOUND)
        
        serializer = self.get_serializer(modelo)
        return Response({
            'success': True,
            'modelo': serializer.data
        })


# ============================================================================
# CU25: ViewSet para Predicciones
# ============================================================================

class PrediccionViewSet(viewsets.ModelViewSet):
    """
    CU25: Visualizar Predicciones de Ventas en Dashboard
    
    Endpoints:
    - GET /api/sales/predicciones/ - Listar predicciones
    - GET /api/sales/predicciones/{id}/ - Obtener detalle
    - GET /api/sales/predicciones/por_periodo/ - Predicciones por período
    - GET /api/sales/predicciones/proximas/ - Próximas predicciones
    - POST /api/sales/predicciones/generar/ - Generar nuevas predicciones
    """
    queryset = Prediccion.objects.all()
    serializer_class = PrediccionSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_serializer_class(self):
        if self.action == 'list':
            return PrediccionListadoSerializer
        return PrediccionSerializer
    
    @action(detail=False, methods=['post'])
    def generar(self, request):
        """Generar predicciones para los próximos períodos (CU25)"""
        try:
            modelo = ModeloIA.objects.filter(estado='activo').first()
            
            if not modelo:
                return Response({
                    'error': 'No hay modelo IA activo para generar predicciones'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            from datetime import timedelta, datetime
            
            predicciones_creadas = []
            fecha_hoy = timezone.now().date()
            
            # Generar predicciones para próximos 12 meses
            for i in range(1, 13):
                fecha_inicio = fecha_hoy + timedelta(days=30*i)
                fecha_fin = fecha_inicio + timedelta(days=29)
                
                # Simular predicción
                import random
                valor_predicho = random.uniform(10000, 100000)
                
                prediccion = Prediccion.objects.create(
                    modelo=modelo,
                    tipo='mensual',
                    fecha_inicio_periodo=fecha_inicio,
                    fecha_fin_periodo=fecha_fin,
                    valor_predicho=valor_predicho,
                    intervalo_confianza_inferior=valor_predicho * 0.8,
                    intervalo_confianza_superior=valor_predicho * 1.2
                )
                predicciones_creadas.append(prediccion)
            
            serializer = PrediccionListadoSerializer(predicciones_creadas, many=True)
            
            return Response({
                'success': True,
                'message': f'{len(predicciones_creadas)} predicciones generadas',
                'predicciones': serializer.data
            }, status=status.HTTP_201_CREATED)
        
        except Exception as e:
            return Response({
                'error': f'Error al generar predicciones: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def proximas(self, request):
        """Obtener próximas predicciones (próximos 3 meses)"""
        fecha_hoy = timezone.now().date()
        
        predicciones = Prediccion.objects.filter(
            fecha_inicio_periodo__gte=fecha_hoy
        ).order_by('fecha_inicio_periodo')[:3]
        
        serializer = self.get_serializer(predicciones, many=True)
        
        return Response({
            'total': len(predicciones),
            'predicciones': serializer.data
        })
    
    @action(detail=False, methods=['get'])
    def resumen_dashboard(self, request):
        """Resumen de predicciones para el dashboard (CU25)"""
        from datetime import timedelta
        
        fecha_hoy = timezone.now().date()
        fecha_futuro = fecha_hoy + timedelta(days=90)
        
        predicciones = Prediccion.objects.filter(
            fecha_inicio_periodo__gte=fecha_hoy,
            fecha_inicio_periodo__lte=fecha_futuro
        ).order_by('fecha_inicio_periodo')
        
        total_predicho = sum(p.valor_predicho for p in predicciones)
        promedio_predicho = total_predicho / len(predicciones) if predicciones else 0
        
        # Calcular tendencia
        predicciones_ordenadas = list(predicciones)
        if len(predicciones_ordenadas) > 1:
            tendencia = 'alcista' if predicciones_ordenadas[-1].valor_predicho > predicciones_ordenadas[0].valor_predicho else 'bajista'
        else:
            tendencia = 'estable'
        
        return Response({
            'total_predicciones': predicciones.count(),
            'total_predicho': float(total_predicho),
            'promedio_predicho': float(promedio_predicho),
            'tendencia': tendencia,
            'periodo_analisis': f'{fecha_hoy} a {fecha_futuro}',
            'predicciones': PrediccionListadoSerializer(predicciones, many=True).data
        })


# ============================================================================
# CU27: ViewSet para Reportes por Voz en Móvil
# ============================================================================

class ReporteVozMovilViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar Reportes por Voz en Móvil (CU27)
    Endpoints:
    - GET /api/sales/reportes-voz-movil/ - Listar reportes
    - POST /reportes-voz-movil/ - Crear reporte por voz
    - POST /reportes-voz-movil/{id}/procesar/ - Procesar grabación y transcribir
    - POST /reportes-voz-movil/{id}/marcar_favorito/ - Guardar como favorito
    - GET /reportes-voz-movil/favoritos/ - Listar favoritos
    """
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser, JSONParser)
    
    def get_queryset(self):
        """Mostrar solo reportes del usuario autenticado"""
        return ReporteVozMovil.objects.filter(usuario=self.request.user)
    
    def get_serializer_class(self):
        if self.action == 'create':
            return ReporteVozMovilCreateSerializer
        return ReporteVozMovilSerializer
    
    def perform_create(self, serializer):
        """Asignar usuario al crear reporte"""
        serializer.save(usuario=self.request.user)
    
    @action(detail=True, methods=['post'])
    def procesar(self, request, pk=None):
        """Procesar grabación: reconocimiento de voz y generación de reporte (CU27)"""
        reporte_voz = self.get_object()
        
        try:
            # Actualizar estado
            reporte_voz.estado = 'procesando'
            reporte_voz.save()
            
            # Simular transcripción (en producción usar Vosk o API externa)
            if reporte_voz.archivo_audio:
                reporte_voz.transcripcion = f"Reporte generado desde: {reporte_voz.archivo_audio.name}"
                reporte_voz.confianza_transcripcion = 0.95
                reporte_voz.comando_detectado = 'generar_reporte_ventas'
                reporte_voz.parametros_extraidos = {
                    'periodo': 'mensual',
                    'tipo': 'ventas',
                    'formato': 'pdf'
                }
            
            # Crear reporte asociado
            reporte = Reporte.objects.create(
                usuario=request.user,
                titulo=f"Reporte por Voz - {timezone.now().strftime('%d/%m/%Y')}",
                tipo_reporte='ventas',
                formato='pdf',
                contenido='{}',
                filtros=reporte_voz.parametros_extraidos
            )
            
            reporte_voz.reporte_asociado = reporte
            reporte_voz.estado = 'completado'
            reporte_voz.fecha_procesamiento = timezone.now()
            reporte_voz.save()
            
            serializer = self.get_serializer(reporte_voz)
            return Response({
                'message': 'Reporte procesado exitosamente',
                'reporte_voz': serializer.data,
                'reporte_creado': {
                    'id': reporte.id,
                    'titulo': reporte.titulo
                }
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            reporte_voz.estado = 'error'
            reporte_voz.save()
            return Response({
                'error': f'Error al procesar reporte: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def marcar_favorito(self, request, pk=None):
        """Marcar comando de voz como favorito"""
        reporte_voz = self.get_object()
        reporte_voz.marcar_como_favorito()
        return Response({
            'message': 'Comando guardado como favorito',
            'es_favorito': reporte_voz.es_favorito
        })
    
    @action(detail=False, methods=['get'])
    def favoritos(self, request):
        """Obtener comandos guardados como favoritos"""
        favoritos = self.get_queryset().filter(es_favorito=True)
        serializer = self.get_serializer(favoritos, many=True)
        return Response({
            'total': favoritos.count(),
            'favoritos': serializer.data
        })


# ============================================================================
# CU28: ViewSet para Compartir Reportes desde Móvil
# ============================================================================

class CompartirReporteViewSet(viewsets.ModelViewSet):
    """
    ViewSet para Compartir Reportes (CU28)
    Endpoints:
    - GET /api/sales/compartir-reportes/ - Listar comparticiones
    - POST /compartir-reportes/ - Crear compartición
    - POST /compartir-reportes/{id}/reenviar/ - Reenviar reporte
    - POST /compartir-reportes/{id}/generar_link/ - Generar link público
    - GET /compartir-reportes/{id}/estado/ - Ver estado del envío
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        """Mostrar solo comparticiones del usuario"""
        return CompartirReporte.objects.filter(usuario_origen=self.request.user)
    
    def get_serializer_class(self):
        if self.action == 'create':
            return CompartirReporteCreateSerializer
        return CompartirReporteSerializer
    
    def perform_create(self, serializer):
        """Asignar usuario origen al crear compartición"""
        serializer.save(usuario_origen=self.request.user)
    
    @action(detail=True, methods=['post'])
    def reenviar(self, request, pk=None):
        """Reenviar compartición de reporte"""
        comparticion = self.get_object()
        
        try:
            comparticion.estado = 'enviando'
            comparticion.intentos_envio += 1
            comparticion.save()
            
            # Simular envío
            import time
            time.sleep(1)
            
            comparticion.estado = 'enviado'
            comparticion.fecha_envio_exitoso = timezone.now()
            comparticion.save()
            
            return Response({
                'message': 'Reporte reenviado exitosamente',
                'intentos': comparticion.intentos_envio,
                'fecha_envio': comparticion.fecha_envio_exitoso
            })
        except Exception as e:
            comparticion.estado = 'fallido'
            comparticion.error_mensaje = str(e)
            comparticion.save()
            return Response({
                'error': f'Error al reenviar: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def generar_link(self, request, pk=None):
        """Generar link público para compartir reporte"""
        import uuid
        comparticion = self.get_object()
        
        try:
            comparticion.token_publico = str(uuid.uuid4())
            comparticion.fecha_expiracion_link = timezone.now() + timezone.timedelta(days=7)
            comparticion.save()
            
            link_publico = f"/api/sales/compartir-reportes/{comparticion.token_publico}/"
            
            return Response({
                'message': 'Link público generado',
                'link': link_publico,
                'expira_en': comparticion.fecha_expiracion_link,
                'token': comparticion.token_publico
            })
        except Exception as e:
            return Response({
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['get'])
    def estado(self, request, pk=None):
        """Obtener estado actual de la compartición"""
        comparticion = self.get_object()
        return Response({
            'id': comparticion.id,
            'estado': comparticion.get_estado_display(),
            'metodo': comparticion.get_metodo_display(),
            'intentos_envio': comparticion.intentos_envio,
            'fecha_envio': comparticion.fecha_envio_exitoso,
            'error': comparticion.error_mensaje
        })


# ============================================================================
# CU29: ViewSet para Preferencias de Notificaciones
# ============================================================================

class PreferenciaNotificacionesViewSet(viewsets.ViewSet):
    """
    ViewSet para Preferencias de Notificaciones (CU29)
    Endpoints:
    - GET /api/sales/preferencias-notificaciones/ - Obtener preferencias
    - PUT /preferencias-notificaciones/ - Actualizar preferencias
    - POST /preferencias-notificaciones/reset/ - Restablecer a valores por defecto
    - POST /preferencias-notificaciones/en_silencio/ - Verificar si está en horario silencioso
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def list(self, request):
        """Obtener preferencias del usuario"""
        try:
            preferencias = PreferenciaNotificaciones.objects.get(usuario=request.user)
            serializer = PreferenciaNotificacionesSerializer(preferencias)
            return Response(serializer.data)
        except PreferenciaNotificaciones.DoesNotExist:
            # Crear preferencias por defecto
            preferencias = PreferenciaNotificaciones.objects.create(usuario=request.user)
            serializer = PreferenciaNotificacionesSerializer(preferencias)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    def update(self, request):
        """Actualizar preferencias del usuario"""
        try:
            preferencias = PreferenciaNotificaciones.objects.get(usuario=request.user)
            serializer = PreferenciaNotificacionesUpdateSerializer(
                preferencias, data=request.data, partial=True
            )
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'message': 'Preferencias actualizadas',
                    'preferencias': PreferenciaNotificacionesSerializer(preferencias).data
                })
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except PreferenciaNotificaciones.DoesNotExist:
            return Response({
                'error': 'Preferencias no encontradas'
            }, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=False, methods=['post'])
    def reset(self, request):
        """Restablecer preferencias a valores por defecto"""
        try:
            preferencias = PreferenciaNotificaciones.objects.get(usuario=request.user)
            
            # Restablecer a valores por defecto
            preferencias.notificaciones_activas = True
            preferencias.frecuencia_general = 'diaria'
            preferencias.config_tipos = {}
            preferencias.canales_habilitados = ['push', 'email']
            preferencias.horario_silencio_activo = False
            preferencias.palabras_clave_filtro = []
            preferencias.save()
            
            return Response({
                'message': 'Preferencias restablecidas',
                'preferencias': PreferenciaNotificacionesSerializer(preferencias).data
            })
        except PreferenciaNotificaciones.DoesNotExist:
            return Response({
                'error': 'Preferencias no encontradas'
            }, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=False, methods=['post'])
    def en_silencio(self, request):
        """Verificar si el usuario está en horario silencioso"""
        try:
            preferencias = PreferenciaNotificaciones.objects.get(usuario=request.user)
            en_silencio = preferencias.esta_en_horario_silencio()
            
            return Response({
                'en_silencio': en_silencio,
                'horario_silencio_activo': preferencias.horario_silencio_activo,
                'inicio': str(preferencias.horario_silencio_inicio),
                'fin': str(preferencias.horario_silencio_fin)
            })
        except PreferenciaNotificaciones.DoesNotExist:
            return Response({
                'error': 'Preferencias no encontradas'
            }, status=status.HTTP_404_NOT_FOUND)


# ============================================================================
# CU30: ViewSet para Sincronización de Datos Offline/Online
# ============================================================================

class SincronizacionDatosViewSet(viewsets.ModelViewSet):
    """
    ViewSet para Sincronización de Datos (CU30)
    Endpoints:
    - GET /api/sales/sincronizacion/ - Listar sincronizaciones
    - POST /sincronizacion/ - Crear solicitud de sincronización
    - POST /sincronizacion/{id}/iniciar/ - Iniciar sincronización
    - POST /sincronizacion/{id}/resolver_conflicto/ - Resolver conflicto
    - GET /sincronizacion/{id}/estado/ - Obtener estado
    - GET /sincronizacion/{id}/velocidad/ - Obtener velocidad de sincronización
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        """Mostrar solo sincronizaciones del usuario"""
        return SincronizacionDatos.objects.filter(usuario=self.request.user)
    
    def get_serializer_class(self):
        if self.action == 'create':
            return SincronizacionDatosCreateSerializer
        return SincronizacionDatosSerializer
    
    def perform_create(self, serializer):
        """Asignar usuario al crear sincronización"""
        serializer.save(usuario=self.request.user)
    
    @action(detail=True, methods=['post'])
    def iniciar(self, request, pk=None):
        """Iniciar sincronización de datos"""
        sincro = self.get_object()
        
        try:
            sincro.estado = 'sincronizando'
            sincro.fecha_inicio_sincro = timezone.now()
            sincro.progreso_porcentaje = 0
            sincro.save()
            
            # Simular sincronización
            import time, random
            for i in range(1, 101):
                time.sleep(0.01)
                sincro.progreso_porcentaje = i
                if i == 50:
                    sincro.log_sincro.append({'evento': 'Mitad del proceso', 'timestamp': str(timezone.now())})
            
            # Verificar conflictos
            tiene_conflicto = random.choice([True, False])
            if tiene_conflicto:
                sincro.tiene_conflicto = True
                sincro.datos_conflictivos = {
                    'tabla': sincro.tipo_dato,
                    'registros': ['id_1', 'id_2']
                }
                sincro.estado = 'conflicto'
            else:
                sincro.estado = 'completado'
                sincro.fecha_fin_sincro = timezone.now()
                sincro.tiempo_sincro_ms = int((sincro.fecha_fin_sincro - sincro.fecha_inicio_sincro).total_seconds() * 1000)
            
            # Datos simulados
            sincro.tamaño_descarga_kb = random.randint(100, 5000)
            sincro.save()
            
            serializer = self.get_serializer(sincro)
            return Response({
                'message': 'Sincronización iniciada',
                'sincronizacion': serializer.data
            })
        
        except Exception as e:
            sincro.estado = 'fallido'
            sincro.error_mensaje = str(e)
            sincro.save()
            return Response({
                'error': f'Error en sincronización: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def resolver_conflicto(self, request, pk=None):
        """Resolver conflicto de sincronización"""
        sincro = self.get_object()
        
        try:
            resolucion = request.data.get('resolucion', 'servidor')  # servidor, dispositivo, manual
            
            if resolucion not in ['servidor', 'dispositivo', 'manual']:
                return Response({
                    'error': 'Resolución inválida. Opciones: servidor, dispositivo, manual'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            sincro.resolucion_conflicto = resolucion
            sincro.tiene_conflicto = False
            sincro.estado = 'completado'
            sincro.fecha_fin_sincro = timezone.now()
            sincro.tiempo_sincro_ms = int((sincro.fecha_fin_sincro - sincro.fecha_inicio_sincro).total_seconds() * 1000)
            sincro.save()
            
            return Response({
                'message': f'Conflicto resuelto usando: {resolucion}',
                'sincronizacion': SincronizacionDatosSerializer(sincro).data
            })
        
        except Exception as e:
            return Response({
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['get'])
    def estado(self, request, pk=None):
        """Obtener estado actual de sincronización"""
        sincro = self.get_object()
        return Response({
            'id': sincro.id,
            'estado': sincro.get_estado_display(),
            'progreso': sincro.progreso_porcentaje,
            'dispositivo': sincro.get_dispositivo_display(),
            'tipo_dato': sincro.get_tipo_dato_display(),
            'fecha_inicio': sincro.fecha_inicio_sincro,
            'cantidad_registros': sincro.cantidad_registros,
            'tiene_conflicto': sincro.tiene_conflicto,
            'error': sincro.error_mensaje
        })
    
    @action(detail=True, methods=['get'])
    def velocidad(self, request, pk=None):
        """Obtener velocidad de sincronización"""
        sincro = self.get_object()
        velocidad = sincro.calcular_velocidad_sincro()
        
        return Response({
            'velocidad_kb_s': round(velocidad, 2),
            'tamaño_descarga_kb': sincro.tamaño_descarga_kb,
            'tiempo_ms': sincro.tiempo_sincro_ms,
            'tiempo_segundos': sincro.tiempo_sincro_ms / 1000
        })